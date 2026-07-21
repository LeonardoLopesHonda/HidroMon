import calendar
import io
import uuid
from datetime import date as date_type
from pathlib import Path

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.db.database import MonitoredItem, Reading

TEMPLATE_PATH = (
    Path(__file__).resolve().parents[3] / "docs" / "templates" / "formulario-monitoramento-captacao.xlsx"
)

MONTH_LABELS_PT = [
    "Jan.", "Fev.", "Mar.", "Abr.", "Mai.", "Jun.",
    "Jul.", "Ago.", "Set.", "Out.", "Nov.", "Dez.",
]  # fmt: skip


def generate_imasul_report(
    db: Session,
    item_id: uuid.UUID,
    year: int,
    tecnico: str,
    crea: str,
    data_filing: date_type,
    observacoes: str | None,
    barramento_durh: str | None,
) -> bytes:
    item = db.query(MonitoredItem).filter(MonitoredItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail=f"MonitoredItem {item_id} not found")
    if item.type != "hidrometro" or not item.durh_number or not item.outorga_number:
        raise HTTPException(
            status_code=422,
            detail="Item não elegível para o relatório IMASUL (requer hidrômetro com DURH e Outorga)",
        )

    readings = (
        db.query(Reading)
        .filter(Reading.item_id == item_id, Reading.valor.isnot(None))
        .order_by(Reading.date, Reading.recorded_at)
        .all()
    )
    monthly = monthly_consumption_by_month(readings, year)

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    ws["D1"] = year
    ws["H5"] = year
    ws["B4"] = f"CAPITAÇÃO\n{item.name}"
    ws["B5"] = item.durh_number
    ws["F5"] = item.outorga_number

    horas_operacao = item.horas_operacao
    for month in range(1, 13):
        vazao_col, tempo_col, periodo_col, row = _cell_for_month(month)
        dias_no_mes = calendar.monthrange(year, month)[1]
        consumo = monthly.get(month, 0.0)
        vazao = consumo / (dias_no_mes * horas_operacao) if horas_operacao else 0.0
        ws[f"{vazao_col}{row}"] = vazao
        ws[f"{tempo_col}{row}"] = horas_operacao
        ws[f"{periodo_col}{row}"] = dias_no_mes

    if observacoes:
        ws["A14"] = _safe_text(observacoes)
    if barramento_durh:
        ws["D24"] = _safe_text(barramento_durh)
    ws["C26"] = _safe_text(tecnico)
    ws["H26"] = _safe_text(crea)
    ws["H28"] = data_filing
    # Template ships with a long-form date format ("segunda-feira, julho..."); override
    # with the short pt-BR convention instead.
    ws["H28"].number_format = "dd/mm/yyyy"

    item.last_tecnico_responsavel = tecnico
    item.last_crea = crea
    db.commit()

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _safe_text(value: str) -> str:
    """Neutralizes formula injection (CWE-1236): a leading =, +, -, or @ would let
    Excel/LibreOffice interpret free-text filing fields as a formula when opened."""
    if value and value[0] in ("=", "+", "-", "@"):
        return f"'{value}"
    return value


def _cell_for_month(month: int) -> tuple[str, str, str, int]:
    """Template packs Jan-Jun into rows 7-12 cols B/C/D and Jul-Dez into the same rows cols F/G/H."""
    row = 6 + (month if month <= 6 else month - 6)
    vazao_col, tempo_col, periodo_col = ("B", "C", "D") if month <= 6 else ("F", "G", "H")
    return vazao_col, tempo_col, periodo_col, row


def monthly_consumption_by_month(readings: list[Reading], year: int) -> dict[int, float]:
    """Monthly consumption (m³) for each of the 12 months of `year`, per the
    IMASUL month-boundary convention (CONTEXT.md -> Formulario de Monitoramento):

        last reading of the month - last reading before the month

    falling back to the month's own first reading when no earlier reading
    exists anywhere in history (start of the item's history) - in which case
    consumption is 0 rather than measured against a baseline that doesn't
    exist. Months with no readings at all are absent from the result (the
    caller treats a missing key as 0, matching the official example's
    zero-filled months).

    `readings` is the item's full history (not just `year`), already sorted
    by (date, recorded_at) ascending, so a January baseline can reach back
    into December of the previous year.

    Mirrors the TS reference implementation: web/src/lib/metrics.ts -> monthlyConsumption()
    (there it computes one month at a time; here compute all 12 at once).
    """
    result: dict[int, float] = {}
    for month in range(1, 13):
        month_start = date_type(year, month, 1)
        month_end = date_type(
            year + (1 if month == 12 else 0),
            1 if month == 12 else month + 1,
            1,
        )
        in_month = [r for r in readings if month_start <= r.date < month_end]
        if not in_month:
            continue

        last_of_month = in_month[-1]
        before = [r for r in readings if r.date < month_start]
        baseline = before[-1] if before else in_month[0]

        result[month] = float(last_of_month.valor) - float(baseline.valor)
    return result
