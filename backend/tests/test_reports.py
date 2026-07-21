import io
import uuid
import zipfile
from datetime import date, datetime, timezone

import openpyxl
import pytest
from fastapi import HTTPException

from app.db.database import Area, MonitoredItem, Reading
from app.services import reports as report_service


def _area(**overrides):
    kwargs = dict(
        id=uuid.uuid4(),
        name="Área de teste",
        frequency="daily",
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc),
    )
    kwargs.update(overrides)
    return Area(**kwargs)


def _item(area_id, **overrides):
    kwargs = dict(
        id=uuid.uuid4(),
        area_id=area_id,
        name="Captação Serraria",
        type="hidrometro",
        limite_outorgado=None,
        unit=None,
        horas_operacao=24,
        corrego_method=None,
        has_horimetro=False,
        disabled=False,
        durh_number="002450",
        outorga_number="3891",
        barramento_durh=None,
        last_tecnico_responsavel=None,
        last_crea=None,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc),
    )
    kwargs.update(overrides)
    return MonitoredItem(**kwargs)


def _reading(item_id, *, d: date, valor: float, **overrides):
    kwargs = dict(
        id=uuid.uuid4(),
        item_id=item_id,
        date=d,
        recorded_at=datetime(d.year, d.month, d.day, 12, tzinfo=timezone.utc),
        valor=valor,
        horimetro=None,
        nivel=None,
        vazao=None,
        raw_values=None,
        observacoes=None,
        created_by=uuid.uuid4(),
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc),
    )
    kwargs.update(overrides)
    return Reading(**kwargs)


@pytest.fixture
def item(db_session):
    area = _area()
    db_session.add(area)
    db_session.flush()
    item = _item(area.id)
    db_session.add(item)
    db_session.flush()
    yield item
    db_session.rollback()
    db_session.query(Reading).filter(Reading.item_id == item.id).delete()
    db_session.query(MonitoredItem).filter(MonitoredItem.id == item.id).delete()
    db_session.query(Area).filter(Area.id == area.id).delete()
    db_session.commit()


# --- pure function: monthly_consumption_by_month -------------------------------------


def test_monthly_consumption_basic_delta():
    item_id = uuid.uuid4()
    readings = [
        _reading(item_id, d=date(2024, 1, 5), valor=100.0),
        _reading(item_id, d=date(2024, 1, 20), valor=150.0),
        _reading(item_id, d=date(2024, 2, 10), valor=200.0),
    ]
    result = report_service.monthly_consumption_by_month(readings, 2024)
    assert result[1] == pytest.approx(150.0 - 100.0)
    assert result[2] == pytest.approx(200.0 - 150.0)


def test_monthly_consumption_zero_fills_empty_month():
    item_id = uuid.uuid4()
    readings = [
        _reading(item_id, d=date(2024, 1, 31), valor=100.0),
        _reading(item_id, d=date(2024, 3, 1), valor=140.0),
    ]
    result = report_service.monthly_consumption_by_month(readings, 2024)
    assert 2 not in result  # no readings in February at all
    assert result[3] == pytest.approx(140.0 - 100.0)


def test_monthly_consumption_start_of_history_has_no_baseline():
    item_id = uuid.uuid4()
    readings = [
        _reading(item_id, d=date(2024, 5, 10), valor=100.0),
        _reading(item_id, d=date(2024, 5, 25), valor=130.0),
    ]
    result = report_service.monthly_consumption_by_month(readings, 2024)
    assert result[5] == pytest.approx(130.0 - 100.0)


def test_monthly_consumption_january_baseline_crosses_year_boundary():
    item_id = uuid.uuid4()
    readings = [
        _reading(item_id, d=date(2023, 12, 28), valor=90.0),
        _reading(item_id, d=date(2024, 1, 15), valor=110.0),
    ]
    result = report_service.monthly_consumption_by_month(readings, 2024)
    assert result[1] == pytest.approx(110.0 - 90.0)


# --- full report generation ------------------------------------------------------------


def test_report_requires_auth(client, item):
    response = client.get(
        "/reports/imasul",
        params={"item_id": str(item.id), "year": 2024, "tecnico": "Fulano", "crea": "123", "data": "2024-01-01"},
    )
    assert response.status_code in (401, 403)


def test_report_rejects_item_without_durh_outorga(db_session):
    area = _area()
    db_session.add(area)
    db_session.flush()
    ineligible = _item(area.id, durh_number=None, outorga_number=None)
    db_session.add(ineligible)
    db_session.flush()

    with pytest.raises(HTTPException) as exc_info:
        report_service.generate_imasul_report(
            db_session, ineligible.id, 2024, "Fulano", "123", date(2024, 1, 1), None, None
        )
    assert exc_info.value.status_code == 422


def test_report_fills_header_and_vazao_leap_year(db_session, item):
    db_session.add_all(
        [
            _reading(item.id, d=date(2024, 1, 5), valor=100.0),
            _reading(item.id, d=date(2024, 1, 31), valor=224.0),  # +124 m3 in January
        ]
    )
    db_session.flush()

    content = report_service.generate_imasul_report(
        db_session,
        item.id,
        2024,
        "Ana Souza",
        "MS-12345",
        date(2024, 12, 20),
        "Sem intercorrências",
        "DURH-BAR-1",
    )

    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # Header fields
    assert ws["D1"].value == 2024
    assert ws["H5"].value == 2024
    assert ws["B4"].value == f"CAPITAÇÃO\n{item.name}"
    assert ws["B5"].value == "002450"
    assert ws["F5"].value == "3891"

    # January: row 7, cols B/C/D (consumption 124 m3 over 31 days, 24h/dia)
    assert ws["B7"].value == pytest.approx(124.0 / (31 * 24))
    assert ws["C7"].value == 24
    assert ws["D7"].value == 31

    # February 2024 is a leap year: 29 calendar days, zero readings -> 0,00
    assert ws["B8"].value == pytest.approx(0.0)
    assert ws["C8"].value == 24
    assert ws["D8"].value == 29

    # Filing fields
    assert ws["A14"].value == "Sem intercorrências"
    assert ws["D24"].value == "DURH-BAR-1"
    assert ws["C26"].value == "Ana Souza"
    assert ws["H26"].value == "MS-12345"
    assert ws["H28"].value == datetime(2024, 12, 20)
    assert ws["H28"].number_format == "dd/mm/yyyy"


def test_report_preserves_template_logos(db_session, item):
    """openpyxl only round-trips embedded images through load_workbook()/save()
    when Pillow is installed; without it, images are silently dropped with no
    error. Guards against that dependency quietly going missing again."""
    content = report_service.generate_imasul_report(
        db_session, item.id, 2025, "Fulano", "123", date(2025, 1, 10), None, None
    )

    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        media_files = [n for n in archive.namelist() if n.startswith("xl/media/")]
    assert len(media_files) == 2


def test_report_persists_tecnico_crea_for_next_prefill(db_session, item):
    report_service.generate_imasul_report(
        db_session, item.id, 2025, "Beatriz Lima", "MS-98765", date(2025, 1, 10), None, None
    )

    db_session.refresh(item)
    assert item.last_tecnico_responsavel == "Beatriz Lima"
    assert item.last_crea == "MS-98765"


def test_report_neutralizes_formula_injection_in_free_text_fields(db_session, item):
    content = report_service.generate_imasul_report(
        db_session,
        item.id,
        2025,
        "=cmd|'/c calc'!A1",
        "+123",
        date(2025, 1, 10),
        "-2+3",
        "@evil",
    )

    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    assert ws["C26"].value == "'=cmd|'/c calc'!A1"
    assert ws["H26"].value == "'+123"
    assert ws["A14"].value == "'-2+3"
    assert ws["D24"].value == "'@evil"


# --- readings export ---------------------------------------------------------------


def test_readings_export_requires_auth(client, item):
    response = client.get(
        "/reports/readings",
        params={"item_id": str(item.id), "from": "2024-01-01", "to": "2024-01-31"},
    )
    assert response.status_code in (401, 403)


def test_readings_export_missing_item_404(db_session):
    with pytest.raises(HTTPException) as exc_info:
        report_service.generate_readings_export(db_session, uuid.uuid4(), date(2024, 1, 1), date(2024, 1, 31))
    assert exc_info.value.status_code == 404


def test_readings_export_hidrometro_columns_and_range_filter(db_session, item):
    db_session.add_all(
        [
            _reading(item.id, d=date(2024, 1, 5), valor=100.0, observacoes="Normal"),
            _reading(item.id, d=date(2024, 1, 20), valor=150.0),
            _reading(item.id, d=date(2024, 2, 1), valor=999.0),  # outside range — excluded
        ]
    )
    db_session.flush()

    content, item_name = report_service.generate_readings_export(db_session, item.id, date(2024, 1, 1), date(2024, 1, 31))
    assert item_name == item.name

    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["Data", "Leitura (m³)", "Observações"]
    assert [c.value for c in ws[2]] == [datetime(2024, 1, 5), 100.0, "Normal"]
    assert [c.value for c in ws[3]] == [datetime(2024, 1, 20), 150.0, None]
    assert ws.max_row == 3  # header + 2 rows, February reading excluded


def test_readings_export_hidrometro_includes_horimetro_column_when_equipped(db_session):
    area = _area()
    db_session.add(area)
    db_session.flush()
    equipped = _item(area.id, has_horimetro=True)
    db_session.add(equipped)
    db_session.flush()
    db_session.add(_reading(equipped.id, d=date(2024, 1, 5), valor=100.0, horimetro=42.0))
    db_session.flush()

    content, _ = report_service.generate_readings_export(db_session, equipped.id, date(2024, 1, 1), date(2024, 1, 31))
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["Data", "Leitura (m³)", "Horímetro (h)", "Observações"]
    assert [c.value for c in ws[2]] == [datetime(2024, 1, 5), 100.0, 42.0, None]

    db_session.rollback()
    db_session.query(Reading).filter(Reading.item_id == equipped.id).delete()
    db_session.query(MonitoredItem).filter(MonitoredItem.id == equipped.id).delete()
    db_session.query(Area).filter(Area.id == area.id).delete()
    db_session.commit()


def test_readings_export_empty_range_has_headers_no_rows(db_session, item):
    content, _ = report_service.generate_readings_export(db_session, item.id, date(2024, 6, 1), date(2024, 6, 30))
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["Data", "Leitura (m³)", "Observações"]
    assert ws.max_row == 1


def test_readings_export_pluviometro_columns(db_session):
    area = _area()
    db_session.add(area)
    db_session.flush()
    pluv = _item(area.id, type="pluviometro", durh_number=None, outorga_number=None)
    db_session.add(pluv)
    db_session.flush()
    db_session.add(_reading(pluv.id, d=date(2024, 3, 1), valor=12.5))
    db_session.flush()

    content, _ = report_service.generate_readings_export(db_session, pluv.id, date(2024, 3, 1), date(2024, 3, 31))
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["Data", "Leitura (mm)", "Observações"]
    assert [c.value for c in ws[2]] == [datetime(2024, 3, 1), 12.5, None]

    db_session.rollback()
    db_session.query(Reading).filter(Reading.item_id == pluv.id).delete()
    db_session.query(MonitoredItem).filter(MonitoredItem.id == pluv.id).delete()
    db_session.query(Area).filter(Area.id == area.id).delete()
    db_session.commit()


def test_readings_export_corrego_regua_columns(db_session):
    area = _area()
    db_session.add(area)
    db_session.flush()
    corrego = _item(area.id, type="corrego", corrego_method="regua", durh_number=None, outorga_number=None)
    db_session.add(corrego)
    db_session.flush()
    db_session.add(_reading(corrego.id, d=date(2024, 3, 1), valor=None, nivel=0.4, vazao=0.456))
    db_session.flush()

    content, _ = report_service.generate_readings_export(db_session, corrego.id, date(2024, 3, 1), date(2024, 3, 31))
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["Data", "Nível (m)", "Vazão (m³/s)", "Observações"]
    assert [c.value for c in ws[2]] == [datetime(2024, 3, 1), 0.4, 0.456, None]

    db_session.rollback()
    db_session.query(Reading).filter(Reading.item_id == corrego.id).delete()
    db_session.query(MonitoredItem).filter(MonitoredItem.id == corrego.id).delete()
    db_session.query(Area).filter(Area.id == area.id).delete()
    db_session.commit()


def test_readings_export_corrego_tambor_columns(db_session):
    area = _area()
    db_session.add(area)
    db_session.flush()
    corrego = _item(area.id, type="corrego", corrego_method="tambor", durh_number=None, outorga_number=None)
    db_session.add(corrego)
    db_session.flush()
    db_session.add(
        _reading(corrego.id, d=date(2024, 3, 1), valor=None, vazao=0.222, raw_values={"t1": 10.0, "t2": 11.0, "t3": 9.0})
    )
    db_session.flush()

    content, _ = report_service.generate_readings_export(db_session, corrego.id, date(2024, 3, 1), date(2024, 3, 31))
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["Data", "T1 (s)", "T2 (s)", "T3 (s)", "Vazão (m³/s)", "Observações"]
    assert [c.value for c in ws[2]] == [datetime(2024, 3, 1), 10.0, 11.0, 9.0, 0.222, None]

    db_session.rollback()
    db_session.query(Reading).filter(Reading.item_id == corrego.id).delete()
    db_session.query(MonitoredItem).filter(MonitoredItem.id == corrego.id).delete()
    db_session.query(Area).filter(Area.id == area.id).delete()
    db_session.commit()


def test_readings_export_neutralizes_formula_injection_in_observacoes(db_session, item):
    db_session.add(_reading(item.id, d=date(2024, 1, 5), valor=100.0, observacoes="=cmd|'/c calc'!A1"))
    db_session.flush()

    content, _ = report_service.generate_readings_export(db_session, item.id, date(2024, 1, 1), date(2024, 1, 31))
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    assert ws["C2"].value == "'=cmd|'/c calc'!A1"


def test_export_filename_slugifies_item_name_and_range():
    filename = report_service.export_filename("Captação Serraria", date(2024, 1, 1), date(2024, 1, 31))
    assert filename == "leituras-captacao-serraria-2024-01-01-a-2024-01-31.xlsx"
